import json
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
import math

def parse_connected_info(start_json):
    """
    return:
    {
      4: {"LocalHost": "...", "LocalPort": ..., "RemoteHost": "...", "RemotePort": ...},
      6: {...}
    }
    """
    result = {}
    for c in start_json.get("connected", []):
        sid = c["socket"]
        result[sid] = {
            "LocalHost": c["local_host"],
            "LocalPort": c["local_port"],
            "RemoteHost": c["remote_host"],
            "RemotePort": c["remote_port"],
        }
    return result

def iperf_json_to_excel_multi_second(
    json_path: str,
    max_seconds: Optional[int] = None,
) -> pd.DataFrame:
    """
    iperf3 JSON → DataFrame → Excel

    行：时间（秒）
    列：stream_xxx（bits/s）
    """

    # ---------- 1. 读取 JSON ----------
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    intervals = data.get("intervals", [])
    connected = data.get("start", {}).get("connected", [])
    timestamp = data.get("start", {}).get("timestamp", []).get("timesecs")
    dt_jst = datetime.fromtimestamp(timestamp,tz=timezone(timedelta(hours=9)))

    rows = []
    all_stream_ids = set()

    # ---------- 2. 解析 intervals ----------
    for interval in intervals:
        streams = interval.get("streams", [])
        if not streams:
            continue

        # 用第一个 stream 的 start 作为该秒时间
        t = (dt_jst + timedelta(seconds=int(streams[0]["start"]))).strftime("%Y-%m-%d %H:%M:%S")

        if max_seconds is not None and t >= max_seconds:
            break

        row = {"Times(UTC+9)": t}

        for s in streams:
            sid = s["socket"]
            col = f"Stream_{sid}"
            row[col] = math.floor(s["bits_per_second"] + 0.5)
            all_stream_ids.add(sid)

        rows.append(row)

    # ---------- 3. 构建 DataFrame ----------
    df = pd.DataFrame(rows)
    df = df.sort_values("Times(UTC+9)").reset_index(drop=True)

    # 补齐缺失 stream 列
    stream_cols = [f"Stream_{sid}" for sid in sorted(all_stream_ids)]
    df = df[["Times(UTC+9)"] + stream_cols]

    # socket → 信息映射
    conn_map = {}
    for c in connected:
        sid = c["socket"]
        conn_map[f"Stream_{sid}"] = {
            "LocalHost": c["local_host"],
            "LocalPort": c["local_port"],
            "RemoteHost": c["remote_host"],
            "RemotePort": c["remote_port"],
        }

    info_rows = []
    for key in ["LocalHost", "LocalPort", "RemoteHost", "RemotePort"]:
        r = {"Times(UTC+9)": key}
        for col in stream_cols:
            r[col] = conn_map.get(col, {}).get(key, "")
        info_rows.append(r)

    info_df = pd.DataFrame(info_rows)

    # 合并：信息行在前，数据在后
    df = pd.concat([info_df, df], ignore_index=True)
    return df


def write_to_excel(excel_path: str, df: pd.DataFrame):
    print(excel_path)

    # ---------- 4. 写 Excel（pandas） ----------
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=2, header=False, sheet_name="iperf")

    # ---------- 5. 用 openpyxl 美化 ----------
    wb = load_workbook(excel_path)
    ws = wb["iperf"]

    # ===== 样式定义 =====
    center = Alignment(horizontal="center", vertical="center")
    double_bottom = Border(
        bottom=Side(style="double")
    )

    no_border = Border()  # 全部无边框

    # ===== 第一行 =====
    ws.cell(row=1, column=1, value="Times(UTC+9)")
    ws.cell(row=2, column=1, value="Throughput")
    # ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # Stream 表头
    for col_idx, col_name in enumerate(df.columns[1:], start=2):
        ws.cell(row=1, column=col_idx, value=col_name)
        ws.cell(row=2, column=col_idx, value="(bit/s)")

    # 第一行：双下边框
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.border = double_bottom

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=c)
        cell.alignment = center
        cell.border = no_border

    # ===== 列宽 =====
    ws.column_dimensions["A"].width = 21
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ===== 冻结窗格（滚动时表头不动）=====
    ws.freeze_panes = "B7"

    # 所有单元格居中
    for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    wb.save(excel_path)


import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
import os

class IperfAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Iperf Analyzer")

        # 保存选中的文件
        self.selected_files = []

        # 主容器
        main_frame = tk.Frame(self.root, padx=20, pady=20)
        main_frame.grid(sticky="nsew")

        # 自适应最小尺寸
        # self.root.update_idletasks()
        self.root.minsize(main_frame.winfo_reqwidth(),
                     main_frame.winfo_reqheight())

        # 标题 Label
        title_label = tk.Label(
            main_frame,
            text="Iperf Analyzer",
            font=("Arial", 16, "bold")
        )
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))

        # 选择文件按钮
        self.select_btn = tk.Button(
            main_frame,
            text="选择文件",
            width=15,
            command=self.select_files
        )
        self.select_btn.grid(row=1, column=0, padx=(0, 10))

        # 开始按钮
        start_btn = tk.Button(
            main_frame,
            text="开始",
            width=15,
            command=self.start_analysis
        )
        start_btn.grid(row=1, column=1)


    def center_window(self):
        self.root.withdraw()
        self.root.update_idletasks()

        # 窗口尺寸
        win_width = self.root.winfo_reqwidth()
        win_height = self.root.winfo_reqheight()

        # 屏幕尺寸
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # 计算居中坐标
        x = (screen_width - win_width) // 2
        y = (screen_height - win_height) // 2

        # 设置窗口位置
        self.root.geometry(f"{win_width}x{win_height}+{x}+{y}")
        self.root.deiconify()

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="选择文件",
            filetypes=[
                ("JSON 文件", "*.json"),
                ("所有文件", "*.*")
            ]
        )
        if files:
            self.selected_files = list(files)
            self.select_btn.config(text="已选中文件")
            print(self.selected_files)

    def start_analysis(self):
        if not self.selected_files:
            messagebox.showwarning("提示", "请先选择 JSON 文件")
            return

        # ========= 这里执行你的分析脚本 =========
        results = []
        for file in self.selected_files:
            df = iperf_json_to_excel_multi_second(
                json_path= file,
                max_seconds=None,  # 或者 10
            )
            results.append([os.path.basename(file), df])
        # ======================================

        # 执行成功
        messagebox.showinfo("成功", "脚本执行成功")

        # 选择结果保存目录
        base_dir = filedialog.askdirectory(title="选择结果保存路径")
        if not base_dir:
            return

        # 生成时间戳
        timestamp = datetime.now().strftime("%y%m%d%H%M%S")
        result_dir_name = f"IperfAnalyzer_Result_{timestamp}"
        result_dir_path = os.path.join(base_dir, result_dir_name)

        try:
            os.makedirs(result_dir_path, exist_ok=False)
        except Exception as e:
            messagebox.showerror("错误", f"创建文件夹失败：\n{e}")

        try:
            for result in results:
                excel_path = os.path.join(result_dir_path, os.path.splitext(result[0])[0] + f'_{timestamp}.xlsx')
                print(result_dir_path)
                print(excel_path)
                write_to_excel(excel_path, result[1])
            messagebox.showinfo(
                "完成",
                f"结果已保存到文件夹：\n{result_dir_path}"
            )
        except Exception as e:
            messagebox.showerror("错误", f"保存失败：\n{e}")


        # 清空状态
        self.selected_files.clear()
        self.select_btn.config(text="选择文件")
        os.startfile(result_dir_path)



if __name__ == "__main__":
    root = tk.Tk()
    app = IperfAnalyzerApp(root)
    root.resizable(False, False)
    app.center_window()
    root.mainloop()
